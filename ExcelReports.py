import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

df = pd.read_excel(r"C:\Users\lenovo\PycharmProjects\FirstProject\supermarket_sales.xlsx")
#print(df)
df = df[['Gender','Product line','Total']]
print(df)
pivot_table = df.pivot_table(index='Gender',columns='Product line',values='Total',aggfunc='sum')
print(pivot_table)
pivot_table.to_excel('pivot_table.xlsx','Report',startrow=2)

wb = load_workbook(r"C:\Users\lenovo\PycharmProjects\FirstProject\pivot_table.xlsx")
sheet=wb['Report']
min_col = wb.active.min_column
max_col = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

bc = BarChart()

data = Reference(sheet,min_col=min_col+1,min_row=min_row,max_col=max_col,max_row=max_row)
categories = Reference(sheet,min_col=min_col,min_row=min_row+1,max_col=max_col,max_row=max_row)

bc.add_data(data, titles_from_data=True)
bc.set_categories(categories)

sheet.add_chart(bc,"B12")
bc.title= "Sales by Product line"
wb.save("barchart.xlsx")